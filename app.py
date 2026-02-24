import os
import pandas as pd
import tkinter as tk
import customtkinter as ctk
from datetime import datetime
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Configuração visual do CustomTkinter
ctk.set_default_color_theme("dark-blue")

# --- FUNÇÃO DE LOG ---


def salvar_log(mensagem, caminho_base, erro=None):
    timestamp = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    pasta_log = os.path.dirname(caminho_base) if caminho_base else os.getcwd()
    log_path = os.path.join(pasta_log, "log_processamento.txt")
    try:
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(f"[{timestamp}] {mensagem}\n")
            if erro:
                f.write(f"DETALHES DO ERRO:\n{erro}\n")
            f.write("-" * 60 + "\n")
    except Exception as e:
        print(f"Erro ao salvar log: {e}")

# --- LÓGICA DE ESTILIZAÇÃO E SOMATÓRIO ---


def apply_styles(ws):
    """
    Aplica estilos e validações pontuais:
    - Célula em A ou E vazia: pinta A CÉLULA de amarelo.
    - Antepenúltima coluna (final-1) preenchida: pinta A CÉLULA de amarelo.
    - Formatação de data dd/mm/yyyy.
    - Rodapé com somatórios em B, E e última coluna.
    """
    AZUL_BEBE = "83caff"
    VERMELHO = "FF0000"
    CINZA_CLARO = "dddddd"
    AMARELO = "FFFF00"
    PRETO = "000000"

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    align_center = Alignment(
        horizontal='center', vertical='center', wrap_text=True)
    font_bold = Font(bold=True, color=PRETO)
    font_header = Font(bold=True, color=PRETO, size=12)

    orig_max_row = ws.max_row
    max_col = ws.max_column
    antepenult_col_idx = max_col - 1
    footer_row = orig_max_row + 1

    # 1. Medidas e Mesclagem
    ws.row_dimensions[1].height = 85
    ws.column_dimensions['A'].width = 5
    if "A1:E1" not in ws.merged_cells:
        ws.merge_cells("A1:E1")
    if max_col >= 6:
        range_f = f"F1:{get_column_letter(max_col)}1"
        if range_f not in ws.merged_cells:
            ws.merge_cells(range_f)

    # 2. Somatórios no Rodapé
    ws.cell(row=footer_row, column=2).value = "QUANTIDADE A SER LIBERADA POR MÊS"
    ws.cell(row=footer_row, column=5).value = f"=SUM(E3:E{orig_max_row})"
    ws.cell(row=footer_row,
            column=max_col).value = f"=SUM({get_column_letter(max_col)}3:{get_column_letter(max_col)}{orig_max_row})"

    # 3. Estilização e Validação Pontual
    for r in range(1, footer_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = align_center

            # Formatação de Data
            if isinstance(cell.value, datetime):
                cell.number_format = 'DD/MM/YYYY'

            # --- REGRAS DE CABEÇALHO E RODAPÉ ---
            if r == 1:
                cell.font = font_header
                cell.fill = PatternFill(
                    start_color=AZUL_BEBE if c <= 5 else VERMELHO, fill_type="solid")
            elif r == 2:
                cell.font = font_bold
                if c == max_col:
                    cell.fill = PatternFill(
                        start_color=AZUL_BEBE, fill_type="solid")
                elif c <= 5:
                    cell.fill = PatternFill(
                        start_color=CINZA_CLARO, fill_type="solid")
                else:
                    cell.fill = PatternFill(
                        start_color=VERMELHO, fill_type="solid")
            elif r == footer_row:
                cell.font = font_bold
                if c in [2, 5, max_col]:
                    cell.fill = PatternFill(
                        start_color=CINZA_CLARO, fill_type="solid")

            # --- REGRAS DE VALIDAÇÃO (AMARELO PONTUAL) ---
            elif 2 < r < footer_row:
                val = str(cell.value).strip() if cell.value is not None else ""

                # Regra 1: Colunas A (1) ou E (5) vazias
                if c in [1, 5] and val == "":
                    cell.fill = PatternFill(
                        start_color=AMARELO, fill_type="solid")

                # Regra 2: Antepenúltima coluna (final-1) diferente de vazio
                elif c == antepenult_col_idx and val != "":
                    cell.fill = PatternFill(
                        start_color=AMARELO, fill_type="solid")

    # 4. Ajuste de largura
    for c in range(2, max_col + 1):
        max_len = 0
        for r in range(1, footer_row + 1):
            cell_val = ws.cell(row=r, column=c).value
            if cell_val and not str(cell_val).startswith("="):
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[get_column_letter(c)].width = min(max_len + 5, 50)
# --- LÓGICA DE PROCESSAMENTO ---


def process_excel(file_path, status_label, progress_bar, app_instance):
    if not file_path:
        messagebox.showwarning("Aviso", "Nenhum arquivo selecionado!")
        return

    salvar_log("Iniciando processamento com Somatório.", file_path)

    try:
        file_name = os.path.splitext(os.path.basename(file_path))[0]
        output_dir = os.path.join(os.path.dirname(file_path), file_name)
        os.makedirs(output_dir, exist_ok=True)

        status_label.configure(
            text="Status: Analisando dados...", text_color="yellow")
        app_instance.update_idletasks()

        xls = pd.ExcelFile(file_path)
        sheets = xls.sheet_names
        unidades = set()
        headers = {}

        for sheet in sheets:
            df = pd.read_excel(xls, sheet_name=sheet, header=None)
            headers[sheet] = df.iloc[0].tolist()
            df.columns = df.iloc[1]
            df = df[2:].reset_index(drop=True)
            if 'UNIDADE' in df.columns:
                unidades.update(df['UNIDADE'].dropna().unique())

        lista_unidades = list(unidades)
        total = len(lista_unidades)

        for i, unidade in enumerate(lista_unidades):
            progress_bar.set((i + 1) / total)
            status_label.configure(
                text=f"Processando: {unidade} ({i+1}/{total})")
            app_instance.update_idletasks()

            output_file = os.path.join(output_dir, f"{unidade}.xlsx")
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                for sheet in sheets:
                    df = pd.read_excel(xls, sheet_name=sheet, header=1)
                    if 'UNIDADE' in df.columns:
                        df_filtered = df[df['UNIDADE'] == unidade]
                        if not df_filtered.empty:
                            header_df = pd.DataFrame(
                                [df_filtered.columns.tolist()], columns=df_filtered.columns)
                            first_row_df = pd.DataFrame(
                                [headers[sheet]], columns=df_filtered.columns)
                            df_final = pd.concat(
                                [first_row_df, header_df, df_filtered], ignore_index=True)
                            df_final.to_excel(
                                writer, sheet_name=sheet, index=False, header=False)

            wb = load_workbook(output_file)
            for sheet in wb.sheetnames:
                apply_styles(wb[sheet])
            wb.save(output_file)
            salvar_log(
                f"Unidade {unidade} concluída com somatório.", file_path)

        status_label.configure(text="Status: Concluído!", text_color="#2ecc71")
        salvar_log("Sucesso total.", file_path)
        messagebox.showinfo(
            "Sucesso", "Processamento e somatórios concluídos!")
        os.startfile(output_dir)

    except Exception as e:
        salvar_log("ERRO", file_path, str(e))
        messagebox.showerror("Erro", f"Ocorreu um erro: {str(e)}")
        status_label.configure(text="Status: Erro", text_color="red")

# --- INTERFACE (MANTIDA) ---


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("CEAF - Data Processor")
        self.geometry("600x450")
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.main_frame = ctk.CTkFrame(self, corner_radius=15)
        self.main_frame.grid(row=0, column=0, padx=20, pady=20, sticky="nsew")
        self.main_frame.grid_columnconfigure(0, weight=1)
        self.title_label = ctk.CTkLabel(
            self.main_frame, text="📊 Processador de Planilhas", font=ctk.CTkFont(size=24, weight="bold"))
        self.title_label.grid(row=0, column=0, padx=20, pady=(30, 10))
        self.subtitle_label = ctk.CTkLabel(
            self.main_frame, text="Divisão automática por UNIDADE", font=ctk.CTkFont(size=13), text_color="gray")
        self.subtitle_label.grid(row=1, column=0, padx=20, pady=(0, 20))
        self.entry_path = ctk.CTkEntry(
            self.main_frame, placeholder_text="Caminho do arquivo...", width=400, height=35)
        self.entry_path.grid(row=2, column=0, padx=20, pady=10)
        self.select_btn = ctk.CTkButton(self.main_frame, text="Explorar Arquivos",
                                        command=self.open_file_dialog, fg_color="#34495e", hover_color="#2c3e50")
        self.select_btn.grid(row=3, column=0, padx=20, pady=5)
        self.progress_bar = ctk.CTkProgressBar(self.main_frame, width=400)
        self.progress_bar.grid(row=4, column=0, padx=20, pady=(30, 10))
        self.progress_bar.set(0)
        self.status_label = ctk.CTkLabel(
            self.main_frame, text="Status: Aguardando seleção", font=ctk.CTkFont(size=12))
        self.status_label.grid(row=5, column=0, padx=20, pady=5)
        self.process_btn = ctk.CTkButton(self.main_frame, text="INICIAR PROCESSAMENTO", font=ctk.CTkFont(size=14, weight="bold"), height=45, fg_color="#2980b9", hover_color="#3498db",
                                         command=lambda: process_excel(self.entry_path.get(), self.status_label, self.progress_bar, self))
        self.process_btn.grid(row=6, column=0, padx=20, pady=(20, 30))

    def open_file_dialog(self):
        path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.entry_path.delete(0, tk.END)
            self.entry_path.insert(0, path)
            self.status_label.configure(
                text="Status: Pronto", text_color="#3498db")
            self.progress_bar.set(0)


if __name__ == "__main__":
    App().mainloop()
